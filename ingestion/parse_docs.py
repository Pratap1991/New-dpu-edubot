"""
DPU EduBot — Document Parser
Handles PDF, DOCX, and Excel files uploaded via the Admin panel.
"""

import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def parse_pdf(filepath: str, batch_id: str, doc_type: str) -> list:
    """Extract text from a PDF file and return chunks."""
    try:
        import pdfplumber
        text = ""
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                text += (page.extract_text() or "") + "\n"
        return _make_chunks(text, filepath, batch_id, doc_type)
    except Exception as e:
        print(f"PDF parse error: {e}")
        return []


def parse_docx(filepath: str, batch_id: str, doc_type: str) -> list:
    """Extract text from a DOCX file and return chunks."""
    try:
        from docx import Document
        doc = Document(filepath)
        text = "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
        return _make_chunks(text, filepath, batch_id, doc_type)
    except Exception as e:
        print(f"DOCX parse error: {e}")
        return []


def parse_excel(filepath: str, batch_id: str, doc_type: str) -> list:
    """Extract text from an Excel file and return chunks."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filepath, data_only=True)
        text = ""
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text += f"\n[Sheet: {sheet_name}]\n"
            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join([str(c) for c in row if c is not None])
                if row_text.strip():
                    text += row_text + "\n"
        return _make_chunks(text, filepath, batch_id, doc_type)
    except Exception as e:
        print(f"Excel parse error: {e}")
        return []


def _make_chunks(text: str, source: str, batch_id: str, doc_type: str) -> list:
    """Split text into overlapping chunks for indexing."""
    try:
        from langchain.text_splitter import RecursiveCharacterTextSplitter
        splitter = RecursiveCharacterTextSplitter(chunk_size=400, chunk_overlap=50)
        raw_chunks = splitter.split_text(text)
    except Exception:
        # Fallback simple splitter
        words = text.split()
        raw_chunks = []
        for i in range(0, len(words), 80):
            chunk = " ".join(words[i:i + 100])
            if chunk.strip():
                raw_chunks.append(chunk)

    return [
        {
            "text": c,
            "source": os.path.basename(source),
            "batch_id": batch_id,
            "category": doc_type,
            "layer": "layer_2",
            "tags": [doc_type, batch_id]
        }
        for c in raw_chunks if len(c.strip()) > 80
    ]
